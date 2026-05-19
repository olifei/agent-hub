#!/usr/bin/env python3
"""Generate a test Excel dataset with English column names using real product data."""

import os

try:
    import openpyxl
except ImportError:
    raise SystemExit("pip install openpyxl   # required to create .xlsx files")


def create_test_dataset(output_path: str = "mcp_server/tests/test_dataset.xlsx"):
    """Create a test Excel with 3 real products across 2 category sheets."""
    wb = openpyxl.Workbook()

    HEADERS = ["product_id", "product_name", "country", "language", "image_url", "company_name"]

    # ── Sheet 1: Luggage Bags Cases (2 products) ────────────────────────────
    ws1 = wb.active
    ws1.title = "Luggage Bags Cases"
    ws1.append(HEADERS)
    ws1.append([
        "1600907870863",
        "Customized Label Colorful Net Crochet Reusable Shopping Cotton Mesh Bag French Grocery Tote Fashionable Lady Farmers Bag",
        "United States",
        "English",
        "https://s.alicdn.com/@sc04/kf/H032f9dfef6ba4d12b6bce93c52d0bd70E.jpg?avif=close&webp=close",
        "Cymbal Shop",
    ])
    ws1.append([
        "1601488769917",
        "High Quality Lip Gloss PU Leather Keychain Bag Private Logo PU Mini Material Lipstick Lipgloss Bags",
        "United States",
        "English",
        "https://s.alicdn.com/@sc04/kf/H4f9fd69ba8c0442e88a857bb40f00447n.jpg?avif=close&webp=close",
        "Cymbal Shop",
    ])

    # ── Sheet 2: Mother Kids Toys (1 product) ───────────────────────────────
    ws2 = wb.create_sheet("Mother Kids Toys")
    ws2.append(HEADERS)
    ws2.append([
        "1601524945369",
        "Living Room Sofa Decoration Sleeping Cushion Stuffed Dolls Cute Sun and Moon Soft Plush Toy Pillows",
        "United States",
        "English",
        "https://s.alicdn.com/@sc04/kf/H931019eb84b04f38af0d26a531b0c330k.jpg?avif=close&webp=close",
        "Cymbal Shop",
    ])

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"✅ Created test dataset: {output_path}")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   Luggage Bags Cases: 2 products (1600907870863, 1601488769917)")
    print(f"   Mother Kids Toys:   1 product  (1601524945369)")
    print(f"\n   Columns: {' | '.join(HEADERS)}")


if __name__ == "__main__":
    create_test_dataset()
